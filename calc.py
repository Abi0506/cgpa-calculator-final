import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
from datetime import datetime

# Grade mapping
grade_points_map = {
    "O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C": 5,
    "U": 0, "SA": 0, "WD": 0
}

def calculate_sgpa_cgpa(df):
    # Convert grades to points
    df['GradePoints'] = df['GRADE'].map(grade_points_map)
    df['CURRSEMS'] = pd.to_numeric(df['CURRSEMS'], errors='coerce')
    df['Credits'] = pd.to_numeric(df['Credits'], errors='coerce')

    sgpa_data = {}

    # Calculate SGPA per semester
    for sem in range(1, 11):
        sem_df = df[df['CURRSEMS'] == sem]
        if not sem_df.empty:
            sgpa_series = sem_df.groupby('715521YYYYYY').apply(
                lambda x: round((x['Credits'] * x['GradePoints']).sum() / x['Credits'].sum(), 3)
                if x['Credits'].sum() > 0 else "-"
            )
        else:
            sgpa_series = pd.Series("-", index=df['715521YYYYYY'].unique())

        sgpa_data[f"SEM{sem}_SGPA"] = sgpa_series

    # Calculate CGPA
    cgpa_series = df.groupby('715521YYYYYY').apply(
        lambda x: round((x['Credits'] * x['GradePoints']).sum() / x['Credits'].sum(), 2)
        if x['Credits'].sum() > 0 else "-"
    )

    # Prepare result DataFrame
    result = df[['INSTCODE', '715521YYYYYY', 'STUDNAME', 'BRANNAME']].drop_duplicates().copy()
    for sem in range(1, 11):
        result[f"SEM{sem}_SGPA"] = result['715521YYYYYY'].map(lambda r: sgpa_data[f"SEM{sem}_SGPA"].get(r, "-"))
    result["CGPA"] = result['715521YYYYYY'].map(lambda r: cgpa_series.get(r, "-"))

    # Sort by roll number
    result = result.sort_values(by='715521YYYYYY')

    return result

def format_excel_center(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Center align all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)

def select_file():

    global result_df
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return

    try:
        df = pd.read_excel(file_path, dtype=str)  # Load all as strings to avoid dtype issues
        required_cols = ["INSTCODE", "715521YYYYYY", "STUDNAME", "BRANNAME", "CURRSEMS", "GRADE", "Credits"]

        if not all(col in df.columns for col in required_cols):
            messagebox.showerror("Error", f"Excel must have columns: {', '.join(required_cols)}")
            return

        # 🔹 Cross-INSTCODE duplicate check
        roll_to_inst = df.groupby("715521YYYYYY")["INSTCODE"].nunique()
        problem_rolls = roll_to_inst[roll_to_inst > 1].index.tolist()

        if problem_rolls:
            dup_rows = df[df["715521YYYYYY"].isin(problem_rolls)][["INSTCODE", "715521YYYYYY"]].drop_duplicates()
            dup_list = "\n".join(f"Roll No: {rn} found in INSTCODE(s): {', '.join(map(str, df[df['715521YYYYYY'] == rn]['INSTCODE'].unique()))}"
                                 for rn in problem_rolls)
            messagebox.showerror(
                "Cross-INSTCODE Duplicate Error",
                f"The following roll numbers appear in more than one INSTCODE:\n\n{dup_list}"
            )
            return

        # Convert numeric columns where needed
        df["CURRSEMS"] = pd.to_numeric(df["CURRSEMS"], errors="coerce")
        df["Credits"] = pd.to_numeric(df["Credits"], errors="coerce")

        result_df = calculate_sgpa_cgpa(df)
        show_table(result_df)
        btn_download.config(state="normal")  # Enable download button
        messagebox.showinfo("Success", "SGPA & CGPA calculated!\nClick 'Download Excel' to save the file.")

    except Exception as e:
        messagebox.showerror("Error", str(e))
def download_excel():
    global result_df
    if result_df is None:
        messagebox.showerror("Error", "No data to download. Please select and process a file first.")
        return

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialfile="SGPA_CGPA_Output.xlsx"
    )
    if not save_path:
        return

    try:
        result_df.to_excel(save_path, index=False)
        format_excel_center(save_path)
        messagebox.showinfo("Success", f"File saved to: {save_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))


def show_table(df):
    for widget in frame_table.winfo_children():
        widget.destroy()

    tree = ttk.Treeview(frame_table, columns=list(df.columns), show='headings')
    tree.pack(fill='both', expand=True)

    for col in df.columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=100)

    for _, row in df.iterrows():
        tree.insert("", "end", values=list(row))


root = tk.Tk()
root.title("SGPA & CGPA Calculator")
root.geometry("900x600")  # Not full screen

btn_select = tk.Button(root, text="Select Excel File", command=select_file)
btn_select.pack(pady=10)

btn_download = tk.Button(root, text="Download Excel", command=download_excel, state="disabled")
btn_download.pack(pady=5)

frame_table = tk.Frame(root)
frame_table.pack(fill='both', expand=True)

result_df = None  # Initialize global variable

root.mainloop()
